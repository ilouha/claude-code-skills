#!/usr/bin/env python3
"""
Material Breakdown by CSI MasterFormat Division
Reads the spatial hierarchy JSON and exports an Excel workbook with:
  Sheet 1: Full material breakdown with CSI classification
  Sheet 2: Division summary
  Sheet 3: Storey x Division matrix
"""

import json
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_JSON = sys.argv[1]
OUTPUT_XLSX = sys.argv[2]

# --- Load JSON ---
with open(INPUT_JSON, "r", encoding="utf-8") as f:
    data = json.load(f)

# --- CSI MasterFormat Division mapping ---
CSI_RULES = [
    # Division 03 - Concrete
    (["concrete", "cast-in-place", "gypcrete"], "03", "Concrete", "03 30 00", "Cast-in-Place Concrete"),
    # Division 04 - Masonry
    (["brick", "masonry", "mortar", "cmu"], "04", "Masonry", "04 21 00", "Clay Unit Masonry"),
    # Division 05 - Metals
    (["steel", "stainless steel", "metal", "aluminum", "bronze", "chrome", "nickel", "iron", "galvanized", "brass"], "05", "Metals", "05 50 00", "Metal Fabrications"),
    # Division 06 - Wood, Plastics & Composites
    (["wood", "lumber", "softwood", "plywood", "cherry", "pine", "oak", "timber", "truss", "glulam", "cabinets", "counter top", "tounge and grove", "ipe deck", "casework"], "06", "Wood, Plastics & Composites", "06 10 00", "Rough Carpentry"),
    # Division 07 - Thermal & Moisture Protection
    (["roofing", "shingle", "insulation", "vapor barrier", "air barrier", "waterproof", "membrane", "protection board", "densglass", "drainage matt", "neoprene", "gasket", "default roof"], "07", "Thermal & Moisture Protection", "07 00 00", "Thermal & Moisture Protection"),
    # Division 08 - Openings
    (["door - frame", "door - panel", "interior door", "window frame", "window screen", "glass - milgard", "glass-marvin", "glazing"], "08", "Openings", "08 00 00", "Openings"),
    # Division 09 - Finishes
    (["gypsum", "plaster", "paint", "epoxy flooring", "tile", "porcelain", "flooring", "finish", "resilient channel", "fabric", "cement board", "color rgb", "sgb"], "09", "Finishes", "09 00 00", "Finishes"),
    # Division 10 - Specialties
    (["access door"], "10", "Specialties", "10 00 00", "Specialties"),
    # Division 11 - Equipment
    (["appliance", "grill", "fridge", "fisher & paykel", "kenmore", "lg electronics", "modern flames", "wildfire", "kohler", "blanco"], "11", "Equipment", "11 40 00", "Residential Equipment"),
    # Division 12 - Furnishings
    (["furnish", "furniture"], "12", "Furnishings", "12 00 00", "Furnishings"),
    # Division 22 - Plumbing
    (["plumbing", "brizo", "waterworks", "house of rohl", "drain", "shower", "toilet", "pipe"], "22", "Plumbing", "22 40 00", "Plumbing Fixtures"),
    # Division 26 - Electrical
    (["led", "electric", "lighting", "outlet", "switch", "plastic, opaque white", "housing"], "26", "Electrical", "26 50 00", "Lighting"),
    # Division 31 - Earthwork
    (["earth", "sand", "gravel", "soil"], "31", "Earthwork", "31 00 00", "Earthwork"),
    # Division 32 - Exterior Improvements
    (["ipe decking"], "32", "Exterior Improvements", "32 18 00", "Site Decking"),
    # Glass (general) - catch-all for "Glass" not caught above
    (["glass"], "08", "Openings", "08 80 00", "Glazing"),
]


def classify_material(mat_name):
    """Classify a material name into CSI MasterFormat."""
    lower = mat_name.lower()
    for patterns, div_num, div_title, section, section_title in CSI_RULES:
        for pat in patterns:
            if pat in lower:
                return div_num, div_title, section, section_title
    return "01", "General Requirements", "01 00 00", "Unclassified"


# --- Collect material usage from all elements ---
material_data = defaultdict(lambda: {
    "element_count": 0,
    "element_types": defaultdict(int),
    "storeys": defaultdict(int),
    "sample_elements": [],
})


def process_elements(elements, storey_name="Unassigned"):
    for elem in elements:
        mats = elem.get("materials", [])
        ifc_type = elem.get("ifc_type", "Unknown")
        elem_name = elem.get("name", "unnamed")
        for mat in mats:
            md = material_data[mat]
            md["element_count"] += 1
            md["element_types"][ifc_type] += 1
            md["storeys"][storey_name] += 1
            if len(md["sample_elements"]) < 3:
                md["sample_elements"].append(elem_name)


# Walk the spatial hierarchy
for site in data["project"]["sites"]:
    for building in site["buildings"]:
        for storey in building["storeys"]:
            storey_name = storey["name"]
            for ifc_type, elems in storey.get("elements_by_type", {}).items():
                process_elements(elems, storey_name)

# Unassigned
if "unassigned_elements" in data:
    for ifc_type, elems in data["unassigned_elements"].get("elements_by_type", {}).items():
        process_elements(elems, "Unassigned")

# --- Build rows ---
rows = []
for mat_name, info in sorted(material_data.items()):
    div_num, div_title, section, section_title = classify_material(mat_name)
    elem_types_str = ", ".join(
        f"{t} ({c})" for t, c in sorted(info["element_types"].items(), key=lambda x: -x[1])
    )
    storeys_str = ", ".join(
        f"{s} ({c})" for s, c in sorted(info["storeys"].items(), key=lambda x: -x[1])
    )
    samples = "; ".join(info["sample_elements"][:3])

    rows.append({
        "div_num": div_num,
        "div_title": div_title,
        "section": section,
        "section_title": section_title,
        "material": mat_name,
        "element_count": info["element_count"],
        "element_types": elem_types_str,
        "storeys": storeys_str,
        "samples": samples,
    })

rows.sort(key=lambda r: (r["div_num"], r["material"]))

# --- Write Excel ---
wb = Workbook()

# ============ SHEET 1: Full Material Breakdown ============
ws = wb.active
ws.title = "Material Breakdown"

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
div_font = Font(name="Calibri", bold=True, size=11)
div_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
normal_font = Font(name="Calibri", size=10)
thin_border = Border(bottom=Side(style="thin", color="D9D9D9"))

headers = [
    "CSI Division", "Division Title", "Section", "Section Title",
    "Material Name", "Element Count", "Element Types", "Storeys Used", "Sample Elements",
]
col_widths = [13, 30, 12, 35, 45, 14, 45, 45, 55]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 30
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
ws.freeze_panes = "A2"

current_div = None
row_idx = 2
for r in rows:
    if r["div_num"] != current_div:
        current_div = r["div_num"]
        div_label = f"Division {r['div_num']} - {r['div_title']}"
        cell = ws.cell(row=row_idx, column=1, value=div_label)
        cell.font = div_font
        cell.fill = div_fill
        for c in range(2, len(headers) + 1):
            ws.cell(row=row_idx, column=c).fill = div_fill
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        row_idx += 1

    values = [
        r["div_num"], r["div_title"], r["section"], r["section_title"],
        r["material"], r["element_count"], r["element_types"],
        r["storeys"], r["samples"],
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if col_idx == 6:
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    row_idx += 1

# ============ SHEET 2: Division Summary ============
ws2 = wb.create_sheet("Division Summary")

div_summary = defaultdict(lambda: {"materials": 0, "elements": 0, "mat_list": []})
for r in rows:
    key = f"{r['div_num']} - {r['div_title']}"
    div_summary[key]["materials"] += 1
    div_summary[key]["elements"] += r["element_count"]
    div_summary[key]["mat_list"].append(r["material"])

sum_headers = ["CSI Division", "Material Count", "Element Count", "Materials"]
sum_widths = [35, 16, 16, 90]

for col_idx, (header, width) in enumerate(zip(sum_headers, sum_widths), 1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

ws2.row_dimensions[1].height = 25
ws2.freeze_panes = "A2"

row_idx = 2
for div_key in sorted(div_summary.keys()):
    info = div_summary[div_key]
    ws2.cell(row=row_idx, column=1, value=div_key).font = Font(name="Calibri", bold=True, size=10)
    ws2.cell(row=row_idx, column=2, value=info["materials"]).alignment = Alignment(horizontal="center")
    ws2.cell(row=row_idx, column=3, value=info["elements"]).alignment = Alignment(horizontal="center")
    mat_cell = ws2.cell(row=row_idx, column=4, value=", ".join(info["mat_list"]))
    mat_cell.font = normal_font
    mat_cell.alignment = Alignment(wrap_text=True)
    row_idx += 1

# Total row
ws2.cell(row=row_idx, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
total_mats = sum(d["materials"] for d in div_summary.values())
total_elems = sum(d["elements"] for d in div_summary.values())
ws2.cell(row=row_idx, column=2, value=total_mats).font = Font(name="Calibri", bold=True, size=11)
ws2.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
ws2.cell(row=row_idx, column=3, value=total_elems).font = Font(name="Calibri", bold=True, size=11)
ws2.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")

# ============ SHEET 3: Storey x Division Matrix ============
ws3 = wb.create_sheet("Storey x Division Matrix")

all_storeys_ordered = []
for site in data["project"]["sites"]:
    for building in site["buildings"]:
        for storey in building["storeys"]:
            all_storeys_ordered.append(storey["name"])
all_storeys_ordered.append("Unassigned")

all_divs = sorted(set(f"Div {r['div_num']}" for r in rows))

matrix = defaultdict(lambda: defaultdict(int))
for mat_name, info in material_data.items():
    d, _, _, _ = classify_material(mat_name)
    dl = f"Div {d}"
    for storey_name, count in info["storeys"].items():
        matrix[storey_name][dl] += count

ws3.cell(row=1, column=1, value="Storey").font = header_font
ws3.cell(row=1, column=1).fill = header_fill
ws3.column_dimensions["A"].width = 35
for col_idx, div in enumerate(all_divs, 2):
    cell = ws3.cell(row=1, column=col_idx, value=div)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    ws3.column_dimensions[get_column_letter(col_idx)].width = 12

total_col = len(all_divs) + 2
ws3.cell(row=1, column=total_col, value="Total").font = header_font
ws3.cell(row=1, column=total_col).fill = header_fill
ws3.cell(row=1, column=total_col).alignment = Alignment(horizontal="center")
ws3.column_dimensions[get_column_letter(total_col)].width = 12
ws3.freeze_panes = "B2"

row_idx = 2
for storey in all_storeys_ordered:
    ws3.cell(row=row_idx, column=1, value=storey).font = Font(name="Calibri", bold=True, size=10)
    row_total = 0
    for col_idx, div in enumerate(all_divs, 2):
        val = matrix[storey][div]
        row_total += val
        cell = ws3.cell(row=row_idx, column=col_idx, value=val if val else "")
        cell.alignment = Alignment(horizontal="center")
    ws3.cell(row=row_idx, column=total_col, value=row_total).font = Font(bold=True)
    ws3.cell(row=row_idx, column=total_col).alignment = Alignment(horizontal="center")
    row_idx += 1

# Column totals
ws3.cell(row=row_idx, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
for col_idx, div in enumerate(all_divs, 2):
    col_total = sum(matrix[s][div] for s in all_storeys_ordered)
    ws3.cell(row=row_idx, column=col_idx, value=col_total).font = Font(bold=True)
    ws3.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")
grand_total = sum(
    sum(matrix[s][d] for d in all_divs) for s in all_storeys_ordered
)
ws3.cell(row=row_idx, column=total_col, value=grand_total).font = Font(bold=True)
ws3.cell(row=row_idx, column=total_col).alignment = Alignment(horizontal="center")

# --- Save ---
wb.save(OUTPUT_XLSX)
print(f"Saved: {OUTPUT_XLSX}")
print(f"Materials classified: {len(rows)}")
print(f"Divisions used: {len(div_summary)}")
print("Sheets: Material Breakdown | Division Summary | Storey x Division Matrix")
