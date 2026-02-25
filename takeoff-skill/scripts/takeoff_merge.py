#!/usr/bin/env python3
"""
takeoff_merge.py - Merge Revit schedule exports into structured JSON.

Reads CSV and/or Excel files containing Revit schedule exports,
normalizes categories, groups instances by Revit Type, and outputs
a single structured JSON file suitable for LLM querying.

Dependencies: openpyxl (for .xlsx support), stdlib only otherwise.
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ASSETS_DIR = Path(__file__).resolve().parent.parent / "assets"

KNOWN_HEADER_PATTERNS = re.compile(
    r"^(Mark|Type|Family|Family and Type|Type Name|Count|Length|Width|Height|"
    r"Area|Volume|Level|Base Constraint|Top Constraint|Offset|Description|"
    r"Comments|Assembly Code|Keynote|Model|Manufacturer|Cost|Phase|"
    r"Head Height|Sill Height|Rough Width|Rough Height|Thickness|"
    r"Unconnected Height|Base Offset|Top Offset|Department|Number|Name|"
    r"Perimeter|Unbounded Height|Limit Offset|Fire Rating)$",
    re.IGNORECASE,
)

SUMMARY_ROW_PATTERNS = re.compile(
    r"^(Grand Total|Subtotal|Total|Sum|Count|Average)(\s|:|$)", re.IGNORECASE
)

# Revit subtotal rows: "A: 23", "Basic Roof: 06 Roof 1-HR Assembly: 5", etc.
# Pattern: text ending with ": <number>"
REVIT_SUBTOTAL_RE = re.compile(r"^.+:\s*\d+\s*$")

# Unit suffixes found in Revit schedule exports (SF, CF, LF, etc.)
UNIT_SUFFIX_RE = re.compile(
    r"^(-?\d[\d,]*\.?\d*)\s*"
    r"(SF|CF|LF|CY|EA|VLF|"
    r"sq\s*ft|cu\s*ft|lin\s*ft|"
    r"ft²|ft³)$",
    re.IGNORECASE,
)

IMPERIAL_RE = re.compile(
    r"^\s*"
    r"(?:(\d+)\s*'\s*-?\s*)?"        # optional feet: 5' or 5' -
    r"(?:"
    r"(\d+)"                          # whole inches
    r"(?:\s+(\d+)\s*/\s*(\d+))?"      # optional fraction: 3/4
    r'\s*"'                           # closing quote for inches
    r"|"
    r"(\d+)\s*'"                      # OR feet-only: 12'
    r")"
    r"\s*$"
)

TYPE_COLUMN_CANDIDATES = [
    "Type",
    "Family and Type",
    "Type Name",
    "Family",
]

# Columns whose values are typically imperial dimensions in Revit exports
DIMENSION_HINT_NAMES = {
    "length", "width", "height", "depth", "offset", "thickness",
    "head height", "sill height", "rough width", "rough height",
    "base offset", "top offset", "unconnected height", "limit offset",
    "perimeter", "unbounded height",
}

# ---------------------------------------------------------------------------
# Config loaders
# ---------------------------------------------------------------------------

@lru_cache(maxsize=1)
def load_category_aliases() -> dict:
    path = ASSETS_DIR / "category-aliases.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Strip the _comment key if present
    data.pop("_comment", None)
    return {k.lower().strip(): v for k, v in data.items()}


@lru_cache(maxsize=1)
def load_skill_config() -> dict:
    path = ASSETS_DIR / "takeoff-skill-config.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Category normalization
# ---------------------------------------------------------------------------

def normalize_category_name(raw: str) -> str:
    """Convert a raw schedule/sheet name into a normalized category key.

    Examples:
        "Wall Schedule Export" -> "walls"
        "Doors" -> "doors"
        "Structural Framing Schedule" -> "beams"
    """
    aliases = load_category_aliases()
    cleaned = raw.strip()

    # Strip common prefixes like "DATA - ", "SCHEDULE - ", etc.
    for prefix in ["data - ", "schedule - ", "data_ ", "data_"]:
        if cleaned.lower().startswith(prefix):
            cleaned = cleaned[len(prefix):].strip()
            break

    # Try exact match (case-insensitive)
    key = cleaned.lower()
    if key in aliases:
        return aliases[key]

    # Try progressively shorter prefixes by stripping common suffixes
    for suffix in [" schedule export", " schedule", " export"]:
        if key.endswith(suffix):
            trimmed = key[: -len(suffix)].strip()
            if trimmed in aliases:
                return aliases[trimmed]

    # Substring match: check if any alias key appears at the start of the raw name
    # Sort by length descending to prefer longer (more specific) matches
    for alias_key in sorted(aliases, key=len, reverse=True):
        if key.startswith(alias_key):
            return aliases[alias_key]

    # Also try after stripping " schedule" variants from the key
    for suffix in [" schedule export", " schedule", " export"]:
        stripped = key
        if suffix in key:
            stripped = key.replace(suffix, "").strip()
        for alias_key in sorted(aliases, key=len, reverse=True):
            if stripped.startswith(alias_key) or alias_key.startswith(stripped):
                return aliases[alias_key]

    # Regex cleanup: strip non-alpha, lowercase
    slug = re.sub(r"[^a-z0-9]+", "_", key).strip("_")

    # Simple pluralization if not already plural
    if slug and not slug.endswith("s"):
        slug += "s"

    return slug


# ---------------------------------------------------------------------------
# Header / type detection
# ---------------------------------------------------------------------------

def detect_header_row(rows: list[list], skip_rows: int = 0) -> int | None:
    """Find the first row that looks like a header (3+ cells matching known patterns).

    Scans the first 10 data rows (after skip_rows).
    """
    start = skip_rows
    end = min(start + 10, len(rows))
    for i in range(start, end):
        row = rows[i]
        matches = sum(
            1 for cell in row
            if cell and KNOWN_HEADER_PATTERNS.match(str(cell).strip())
        )
        if matches >= 3:
            return i
    return None


def detect_type_column(headers: list[str]) -> str | None:
    """Find the best column to use for grouping instances by Revit Type.

    Priority order:
    1. Exact match against TYPE_COLUMN_CANDIDATES (case-insensitive)
    2. Any column containing "type" (case-insensitive)
    """
    lower_headers = [h.lower().strip() for h in headers]

    for candidate in TYPE_COLUMN_CANDIDATES:
        cl = candidate.lower()
        if cl in lower_headers:
            idx = lower_headers.index(cl)
            return headers[idx]

    for h in headers:
        if "type" in h.lower():
            return h

    return None


# ---------------------------------------------------------------------------
# Row filtering
# ---------------------------------------------------------------------------

def is_data_row(row: list, num_columns: int) -> bool:
    """Return True if the row contains actual instance data (not blank/summary)."""
    # All cells empty?
    non_empty = [c for c in row[:num_columns] if c is not None and str(c).strip()]
    if not non_empty:
        return False

    # Summary/total row?
    first_cell = str(row[0]).strip() if row else ""
    if SUMMARY_ROW_PATTERNS.match(first_cell):
        return False

    # Revit subtotal row? e.g. "A: 23", "Floor: 03 Conc. Slab 5": 2"
    # These have a count suffix and typically many empty middle columns
    if REVIT_SUBTOTAL_RE.match(first_cell):
        # Confirm: most middle columns are empty (subtotals aggregate only a few)
        empty_count = sum(1 for c in row[:num_columns] if c is None or str(c).strip() == "")
        if empty_count >= num_columns * 0.5:
            return False

    return True


# ---------------------------------------------------------------------------
# Imperial dimension parsing
# ---------------------------------------------------------------------------

def parse_imperial(text: str) -> float | None:
    """Parse a Revit imperial dimension string into decimal feet.

    Handles:
        5' - 3"      -> 5.25
        1'-6 1/2"    -> 1.5417
        3' - 6 3/4"  -> 3.5625
        0' - 11 1/4" -> 0.9375
        10' - 0"     -> 10.0
        8"           -> 0.6667
        12'          -> 12.0

    Returns None if the string is not an imperial dimension.
    """
    if not isinstance(text, str):
        return None

    m = IMPERIAL_RE.match(text)
    if not m:
        return None

    feet_str, inches_str, frac_num_str, frac_den_str, feet_only_str = m.groups()

    # Feet-only pattern (e.g. "12'")
    if feet_only_str is not None:
        return float(feet_only_str)

    feet = int(feet_str) if feet_str else 0
    inches = int(inches_str) if inches_str else 0
    fraction = 0.0
    if frac_num_str and frac_den_str:
        den = int(frac_den_str)
        if den != 0:
            fraction = int(frac_num_str) / den

    total_inches = inches + fraction
    return round(feet + total_inches / 12.0, 4)


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

def coerce_value(raw):
    """Convert a raw cell value to an appropriate Python type.

    Priority:
    1. None/empty -> None
    2. Already numeric -> pass through
    3. "Yes"/"No" -> True/False
    4. Imperial dimension -> float (decimal feet)
    5. Integer pattern -> int
    6. Float pattern (incl. comma thousands) -> float
    7. Fallback -> stripped str
    """
    if raw is None:
        return None

    if isinstance(raw, (int, float)):
        return raw

    text = str(raw).strip()
    if not text:
        return None

    # N/A
    if text.lower() in ("n/a", "na", "-"):
        return None

    # Boolean
    if text.lower() == "yes":
        return True
    if text.lower() == "no":
        return False

    # Imperial dimension (try first, before plain number)
    imp = parse_imperial(text)
    if imp is not None:
        return imp

    # Unit suffix (e.g. "9 SF", "6.21 CF", "1234 LF")
    unit_match = UNIT_SUFFIX_RE.match(text)
    if unit_match:
        num_str = unit_match.group(1).replace(",", "")
        if "." in num_str:
            return float(num_str)
        return int(num_str)

    # Integer
    if re.match(r"^-?\d+$", text):
        return int(text)

    # Float (with optional comma thousands separator)
    cleaned = text.replace(",", "")
    if re.match(r"^-?\d+\.\d+$", cleaned):
        return float(cleaned)

    return text


# ---------------------------------------------------------------------------
# File readers
# ---------------------------------------------------------------------------

def read_csv_file(path: Path) -> list[dict]:
    """Read a CSV file and return a list of sheet dicts.

    Returns: [{"name": <filename_stem>, "rows": [[cell, ...], ...]}]
    """
    rows = []
    # Try utf-8-sig first (handles BOM), fall back to utf-8
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                reader = csv.reader(f)
                rows = [row for row in reader]
            break
        except UnicodeDecodeError:
            continue

    return [{"name": path.stem, "rows": rows}]


def read_excel_file(path: Path) -> list[dict]:
    """Read an Excel file and return a list of sheet dicts.

    Returns: [{"name": <sheet_name>, "rows": [[cell, ...], ...]}, ...]
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required for .xlsx files. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        sheets.append({"name": sheet_name, "rows": rows})
    wb.close()
    return sheets


def read_file(path: Path) -> list[dict]:
    """Dispatch to the correct reader based on file extension."""
    ext = path.suffix.lower()
    if ext == ".csv":
        return read_csv_file(path)
    elif ext in (".xlsx", ".xls"):
        return read_excel_file(path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


# ---------------------------------------------------------------------------
# Sheet processing
# ---------------------------------------------------------------------------

def deduplicate_headers(headers: list[str]) -> list[str]:
    """Append _2, _3, etc. to duplicate column names."""
    seen = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            result.append(h)
    return result


def process_sheet(
    name: str,
    rows: list[list],
    type_column_override: str | None = None,
    skip_rows: int = 0,
) -> dict | None:
    """Process a single sheet into a categorized dict of typed instance groups.

    Returns: {
        "category": str,
        "types": {"TypeName": [{col: val, ...}, ...], ...},
        "instance_count": int,
        "warnings": [str, ...]
    }
    or None if the sheet is empty/unparseable.
    """
    warnings = []

    if not rows or len(rows) < 2:
        warnings.append(f"Sheet '{name}' is empty or has insufficient rows, skipped.")
        return {"category": None, "types": {}, "instance_count": 0, "warnings": warnings}

    # Detect header row
    header_idx = detect_header_row(rows, skip_rows=skip_rows)
    if header_idx is None:
        # Fallback: use the first row after skip_rows as header
        header_idx = skip_rows
        warnings.append(f"Sheet '{name}': no obvious header row detected, using row {header_idx + 1}.")

    raw_headers = [str(c).strip() if c else "" for c in rows[header_idx]]
    headers = deduplicate_headers(raw_headers)
    num_columns = len(headers)

    # Detect type column
    type_col = type_column_override
    if type_col and type_col not in headers:
        warnings.append(
            f"Sheet '{name}': specified type column '{type_col}' not found in headers. Auto-detecting."
        )
        type_col = None

    if not type_col:
        type_col = detect_type_column(headers)

    if not type_col:
        warnings.append(f"Sheet '{name}': no type column found, grouping all under '_untyped'.")

    # Process data rows
    types: dict[str, list[dict]] = {}
    instance_count = 0

    for row in rows[header_idx + 1:]:
        # Pad row to header length
        padded = row + [None] * max(0, num_columns - len(row))

        if not is_data_row(padded, num_columns):
            continue

        # Build instance dict
        instance = {}
        for j, header in enumerate(headers):
            if not header:
                continue
            instance[header] = coerce_value(padded[j])

        # Determine type group
        if type_col and type_col in instance:
            type_name = str(instance[type_col]).strip() if instance[type_col] else "_untyped"
        else:
            type_name = "_untyped"

        if type_name not in types:
            types[type_name] = []
        types[type_name].append(instance)
        instance_count += 1

    if instance_count == 0:
        warnings.append(f"Sheet '{name}': no data rows found after header, skipped.")

    category = normalize_category_name(name)

    return {
        "category": category,
        "types": types,
        "instance_count": instance_count,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Merge orchestrator
# ---------------------------------------------------------------------------

def merge_all(
    input_paths: list[Path],
    output_path: Path,
    type_column_override: str | None = None,
    pretty: bool = False,
    skip_rows: int = 0,
) -> dict:
    """Read all input files, process sheets, merge into output JSON."""
    categories: dict[str, dict[str, list[dict]]] = {}
    all_warnings: list[str] = []
    source_files: list[str] = []
    total_instances = 0

    for path in input_paths:
        if not path.exists():
            all_warnings.append(f"File not found: {path}")
            continue

        source_files.append(path.name)

        try:
            sheets = read_file(path)
        except Exception as e:
            all_warnings.append(f"Error reading {path.name}: {e}")
            continue

        for sheet in sheets:
            result = process_sheet(
                sheet["name"],
                sheet["rows"],
                type_column_override=type_column_override,
                skip_rows=skip_rows,
            )

            if result is None:
                continue

            all_warnings.extend(result["warnings"])

            cat = result["category"]
            if not cat or result["instance_count"] == 0:
                continue

            total_instances += result["instance_count"]

            # Merge type groups into category
            if cat not in categories:
                categories[cat] = {}
            for type_name, instances in result["types"].items():
                if type_name in categories[cat]:
                    categories[cat][type_name].extend(instances)
                else:
                    categories[cat][type_name] = instances

    # Build category counts
    category_counts = {cat: sum(len(v) for v in types.values()) for cat, types in categories.items()}

    # Build output
    output = {
        "summary": {
            "source_files": source_files,
            "categories": category_counts,
            "total_instances": total_instances,
            "warnings": all_warnings,
            "merged_at": datetime.now(timezone.utc).isoformat(),
        },
        "categories": categories,
    }

    # Write JSON
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2 if pretty else None, ensure_ascii=False)

    return output


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def print_summary(output: dict, output_path: Path):
    """Print a human-readable summary to stdout."""
    summary = output["summary"]
    cat_counts = summary["categories"]

    # File size
    size_bytes = output_path.stat().st_size
    if size_bytes > 1_000_000:
        size_str = f"{size_bytes / 1_000_000:.1f} MB"
    else:
        size_str = f"{size_bytes / 1_000:.1f} KB"

    # Category list
    cat_parts = [f"{cat} ({count})" for cat, count in sorted(cat_counts.items())]
    cat_str = ", ".join(cat_parts) if cat_parts else "(none)"

    print()
    print("Takeoff Merge Complete")
    print(f"  Files processed:  {len(summary['source_files'])}")
    print(f"  Categories:       {cat_str}")
    print(f"  Total instances:  {summary['total_instances']}")
    print(f"  Warnings:         {len(summary['warnings'])}")
    print(f"  Output:           {output_path} ({size_str})")

    if summary["warnings"]:
        print()
        print("  Warnings:")
        for w in summary["warnings"]:
            print(f"    - {w}")
    print()


def main():
    parser = argparse.ArgumentParser(
        description="Merge Revit schedule exports (CSV/Excel) into structured JSON.",
        epilog="Example: python takeoff_merge.py output.json Wall_Schedule.csv Doors.xlsx --pretty",
    )
    parser.add_argument(
        "output",
        type=Path,
        help="Output JSON file path",
    )
    parser.add_argument(
        "inputs",
        nargs="+",
        type=Path,
        help="Input CSV or Excel files",
    )
    parser.add_argument(
        "--type-column",
        default=None,
        help="Override the type column name (e.g. 'Family and Type')",
    )
    parser.add_argument(
        "--pretty",
        action="store_true",
        help="Pretty-print the output JSON",
    )
    parser.add_argument(
        "--skip-rows",
        type=int,
        default=0,
        help="Number of leading rows to skip in each sheet (before header detection)",
    )

    args = parser.parse_args()

    output = merge_all(
        input_paths=args.inputs,
        output_path=args.output,
        type_column_override=args.type_column,
        pretty=args.pretty,
        skip_rows=args.skip_rows,
    )

    print_summary(output, args.output)


if __name__ == "__main__":
    main()
